import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from difflib import get_close_matches
from pyautogui import alert



class Dados:
    def __init__(self, arqPlanilha, sede):
        self.arqPlanilha = Path(arqPlanilha)
        
        self.arquivo_progresso = self.arqPlanilha.parent / 'progresso.log'
        self.arquivo_notas = self.arqPlanilha.parent / 'num_notas.txt'

        base_path_matriz = r'base\CPF_matriz.xlsx'
        base_path_filial = r'base\CPF_filial.xlsx'

        if sede == 'Matriz':
            self.base_df = pd.read_excel(base_path_matriz, usecols=["ResponsávelFinanceiro", "CPF"])
        else:
            self.base_df = pd.read_excel(base_path_filial, usecols=["ResponsávelFinanceiro", "CPF"])
        

    def encontrar_melhor_match(self, nome):
        """
        Recebe um nome e retorna uma tupla contendo:
        - o nome da base de dados que mais se aproxima do nome informado;
        - o CPF associado a esse nome.
        Caso nenhum nome seja encontrado, retorna (None, None).
        """
        # Extrai a lista de nomes da coluna "ResponsávelFinanceiro".
        lista_nomes = self.base_df["ResponsávelFinanceiro"].tolist()
        
        # Usa get_close_matches para encontrar o nome mais próximo.
        # n=1: retorna apenas a melhor correspondência.
        # cutoff=0.0: nenhum corte de similaridade; ajuste se necessário.
        correspondencias = get_close_matches(nome, lista_nomes, n=1, cutoff=0.85)
        
        if correspondencias:
            melhor_nome = correspondencias[0]
            # Recupera o CPF associado a esse nome.
            cpf = self.base_df.loc[self.base_df.iloc[:, 0] == melhor_nome].iloc[0, 1]
        else:
            melhor_nome = None
            cpf = None

        return melhor_nome, cpf
    
    
    def formata_planilha(self, arqPlanilha):
        wb = load_workbook(arqPlanilha)

        if not 'dados' in wb.sheetnames:
            self.dados_origem = pd.read_excel(arqPlanilha, 'dados_origem', header=1, skipfooter=1)

            resultado = self.dados_origem["ResponsávelFinanceiro"].apply(
                lambda nome: pd.Series(self.encontrar_melhor_match(nome), index=["ResponsávelFinanceiro", "CPF"])
            )
            self.dados_destino = self.dados_origem.drop(columns=["ResponsávelFinanceiro", "CPF"]).join(resultado)

            dados_faltantes = self.dados_destino['ResponsávelFinanceiro'].isna()
            clientes_novos = pd.DataFrame()
            if dados_faltantes.any():
                clientes_novos = self.dados_origem.loc[dados_faltantes]
                alert(title='Clientes não cadastrados encontrados',
                        text='\n'.join(clientes_novos['ResponsávelFinanceiro'].to_list())
                )
                self.dados_destino.drop(clientes_novos.index, inplace=True)

            with pd.ExcelWriter(arqPlanilha, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                self.dados_destino.to_excel(writer, sheet_name="dados", index=False, startrow=1)
                if not clientes_novos.empty:
                    clientes_novos.to_excel(writer, sheet_name="clientes_novos", index=False)


    def obter_dados(self):
        wb = load_workbook(self.arqPlanilha)

        if 'dados' in wb.sheetnames:
            self.dados = pd.read_excel(self.arqPlanilha, 'dados', header=1)#, skipfooter=1)

            if 'Notas' not in self.dados.columns:
                self.dados['Notas'] = None
            
            self.dados['Aluno'] = self.dados['Aluno'].apply(lambda i: i.split()[0])

            self.dados.loc[self.dados['Turma'].str.contains('Y1|Y2|Year'), 'Acumulador'] = '2'
            self.dados['Acumulador'] = self.dados['Acumulador'].fillna('1')

            self.dados['Mensalidade'] = self.dados['Mensalidade'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))
            self.dados['ValorTotal'] = self.dados['ValorTotal'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))
            self.dados['Alimentação'] = self.dados['Alimentação'].apply(lambda x: '{:0.2f}'.format(x).replace('.',','))

            return self.dados
        else:
            return None
    

    def registra_numero_notas(self, index_df, num_nota):
        wb = load_workbook(self.arqPlanilha)
        sheet = wb['dados']

        # Adicionar a coluna 'Status' se não existir
        if 'Notas' not in [celula.value for celula in sheet[2]]:
            sheet.cell(row=2, column=sheet.max_column + 1).value = 'Notas'

        notas_col_index = [celula.value for celula in sheet[2]].index('Notas') + 1
        sheet.cell(row=index_df+3, column=notas_col_index).value = num_nota

        wb.save(self.arqPlanilha)



if __name__ == '__main__':
    arquivo_planilha = r"C:\Users\novoa\OneDrive\Área de Trabalho\notas_MB\planilhas\zona_norte\escola_canadenseZN_fev25\Maple Bear Fev 25.xlsx"
    dados = Dados(arquivo_planilha, 'Matriz')
    # df = dados.obter_dados()
    # print(df)
    # print(list(df[df['Notas'].isna()].itertuples()))
    # print(list(df.itertuples()))
    dados.formata_planilha(arquivo_planilha)
